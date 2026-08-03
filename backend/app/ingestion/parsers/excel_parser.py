import logging

from openpyxl import load_workbook

from .base_parser import BaseParser, ParsedPage
from .table_utils import rows_to_markdown_chunks

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    def parse(self, file_path: str) -> list[ParsedPage]:
        pages = []

        wb = load_workbook(file_path, data_only=True, read_only=True)
        try:
            for sheet_idx, ws in enumerate(wb.worksheets, start=1):
                rows_iter = ws.iter_rows(values_only=True)

                # Header = dòng KHÔNG rỗng ĐẦU TIÊN. Bỏ qua các dòng trống dẫn đầu —
                # nhiều file export (vd dữ liệu tài chính Simplize) chèn 1+ dòng trống
                # ở đầu rồi mới tới nội dung. TRƯỚC ĐÂY coi cứng dòng đầu là header và
                # bỏ CẢ sheet nếu nó rỗng -> file rơi vào "không trích xuất được nội dung".
                header = None
                for row in rows_iter:
                    cells = [self.clean_text(str(c)) if c is not None else "" for c in row]
                    if any(cells):
                        header = cells
                        break
                if header is None:
                    continue  # sheet rỗng hoàn toàn

                data_rows = [
                    [self.clean_text(str(c)) if c is not None else "" for c in row]
                    for row in rows_iter
                    if any(cell is not None and str(cell).strip() for cell in row)
                ]
                if not data_rows:
                    continue  # chỉ có header, không có dữ liệu

                # Cắt bỏ CỘT trống ở đuôi: sheet thường khai báo nhiều cột nhưng phần lớn
                # rỗng -> chunk đầy "| |" nhiễu, làm loãng embedding số liệu. Bề rộng giữ
                # lại = cột có nội dung xa nhất trong header + mọi data row (không mất dữ liệu).
                width = 0
                for r in (header, *data_rows):
                    for j in range(len(r) - 1, -1, -1):
                        if r[j].strip():
                            if j + 1 > width:
                                width = j + 1
                            break
                if width:
                    header = header[:width]
                    data_rows = [r[:width] for r in data_rows]

                tables = rows_to_markdown_chunks(
                    header, data_rows, start_row_number=2, sheet_name=ws.title
                )

                pages.append(
                    ParsedPage(
                        page_num=sheet_idx,
                        text="",
                        metadata={
                            "tables": tables,
                            "sheet_name": ws.title,
                            "source": "excel",
                            "char_count": sum(len(t) for t in tables),
                        },
                    )
                )
        finally:
            wb.close()

        return pages
